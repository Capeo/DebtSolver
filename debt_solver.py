# Program for finding minimum number of transactions needed to resolve debt matrix.
# Priority is given to minimizing transactions involving users without Vipps

import argparse
import openpyxl as pyxl
import re


def inc_char(text, chlist='ABCDEFGHIJKLMNOPQRSTUVWXYZ', steps=1):
    result = text
    for i in range(steps):
        result = inc_char_once(result, chlist=chlist)
    return result


def inc_char_once(text, chlist):
    # Unique and sort
    chlist = ''.join(sorted(set(str(chlist))))
    chlen = len(chlist)
    if not chlen:
        return ''
    text = str(text)
    # Replace all chars but chlist
    text = re.sub('[^' + chlist + ']', '', text)
    if not len(text):
        return chlist[0]
    # Increment
    inc = ''
    over = False
    for i in range(1, len(text) + 1):
        lchar = text[-i]
        pos = chlist.find(lchar) + 1
        if pos < chlen:
            inc = chlist[pos] + inc
            over = False
            break
        else:
            inc = chlist[0] + inc
            over = True
    if over:
        inc += chlist[0]
    result = text[0:-len(inc)] + inc
    return result


# Sort lenders and borrowers, prioritizing those without Vipps
def sort_debts(loans, debts, has_vipps_dict):
    sorted_loans = sorted(loans, key=lambda x: x[1], reverse=True)
    sorted_debts = sorted(debts, key=lambda x: x[1], reverse=True)

    for loan in sorted_loans:
        lender = loan[0]
        if not has_vipps_dict[lender]:
            sorted_loans.remove(loan)
            sorted_loans.insert(0, loan)

    for debt in sorted_debts:
        debtor = debt[0]
        if not has_vipps_dict[debtor]:
            sorted_debts.remove(debt)
            sorted_debts.insert(0, debt)
    return sorted_loans, sorted_debts


def compute_payments(workbook, nr_participants, input_table_start_cell, result_table_start_cell):
    wb = pyxl.load_workbook(workbook, data_only=True)
    ws = wb.worksheets[0]

    # Read data from xlsm file
    row_index = re.search("\d", input_table_start_cell).start()
    initial_column = input_table_start_cell[:row_index]
    initial_row = int(input_table_start_cell[row_index:]) + 1
    end_row = initial_row + nr_participants - 1

    participants_column = initial_column
    participants_range = ws[participants_column + str(initial_row):participants_column + str(end_row)]
    participants = []
    participants_index = {}
    i = 0
    for cell in participants_range:
        participants.append(cell[0].value)
        participants_index[cell[0].value] = i
        i += 1

    has_vipps_column = inc_char(participants_column, steps=1)
    has_vipps_range = ws[has_vipps_column + str(initial_row):has_vipps_column + str(end_row)]
    has_vipps = []
    for cell in has_vipps_range:
        has_vipps.append(cell[0].value)

    netto_column = inc_char(has_vipps_column, steps=7)
    netto_range = ws[netto_column + str(initial_row):netto_column + str(end_row)]
    netto = []
    for cell in netto_range:
        netto.append(cell[0].value)

    # Define net lenders, borrowers, and who has Vipps
    loans = []
    debts = []
    has_vipps_dict = {}

    for i in range(len(participants)):
        participant = participants[i]
        net_value = round(netto[i])
        if net_value > 0:
            loans.append((participant, net_value))
        elif net_value < 0:
            debts.append((participant, abs(net_value)))
        participant_has_vipps = has_vipps[i]
        if participant_has_vipps == "Ja":
            has_vipps_dict[participant] = True
        else:
            has_vipps_dict[participant] = False

    payments = []

    # compute payments
    while loans:
        loans, debts = sort_debts(loans, debts, has_vipps_dict)
        loan = loans.pop(0)
        debt = debts.pop(0)
        if loan[1] > debt[1]:
            payments.append((debt[0], loan[0], debt[1]))
            print(debt[0], "pays", loan[0], debt[1])
            loans.append((loan[0], loan[1] - debt[1]))
        elif loan[1] < debt[1]:
            payments.append((debt[0], loan[0], loan[1]))
            print(debt[0], "pays", loan[0], loan[1])
            debts.append((debt[0], debt[1] - loan[1]))
        else:
            payments.append((debt[0], loan[0], debt[1]))
            print(debt[0], "pays", loan[0], debt[1])

    # write payments to file
    wb_save = pyxl.load_workbook(workbook, keep_vba=True)
    ws_save = wb_save['Betalingsstruktur']
    row_index = re.search("\d", result_table_start_cell).start()
    first_payment_col = inc_char(result_table_start_cell[:row_index], steps=1)
    first_payment_row = int(result_table_start_cell[row_index:]) + 1
    for payment in payments:
        sender = payment[0]
        receiver = payment[1]
        amount = payment[2]
        sender_index = participants_index[sender]
        receiver_index = participants_index[receiver]
        col = inc_char(first_payment_col, steps=receiver_index)
        row = first_payment_row + sender_index
        ws_save[col + str(row)] = amount
    wb_save.save(workbook)

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Compute payments based on data in excel file and write results to same file")
    parser.add_argument("-wb", nargs='+', help="excel file containing expenses data.")
    parser.add_argument("-p", nargs='?', default=8, help="number of participants. Default 8.", type=int)
    parser.add_argument("-dc", nargs='?', default='B4', help="first cell of data table in excel file. Default B4.")
    parser.add_argument("-rc", nargs='?', default='A1', help="first cell of results table in excel file. Default A1.")
    args = parser.parse_args()
    compute_payments(args.wb[0], args.p, args.dc, args.rc)